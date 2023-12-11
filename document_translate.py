import openpyxl
from openpyxl.styles import Font
from googletrans import Translator
from os import path, makedirs
import csv
import datetime
import re
import settings
import sqlite3

def translate(cell_value):
    return translator.translate(cell_value).text if cell_value is not None else ""

def calculate_amount(cell_value):
    if cell_value == "未":
        return "unknown"
    
    result = int(re.sub(r"\D", "", str(cell_value)))
    if 0 < result < 120:
        return str(result + 5) + "0000"
    elif 119 < result:
        return str(result + 10) + "0000"

def check_cell_value(index_cell, cell, temp):
    if index_cell in [2, 11, 14]:
        temp.append(translate(cell.value))
    if index_cell == 13 and cell.value == "無":
        cell.value = "unknown"
    if index_cell == 12:
        try:
            temp.append(
                translate(
                    openpyxl.utils.datetime.from_excel(cell.value).strftime("%Y")
                )
            )
        except:
            temp.append(translate(cell.value))
    if index_cell == 15:
        temp.append(calculate_amount(cell.value))

def process_worksheet(wb, index, temp_list, amount_list):
    temp = []
    ws = wb.worksheets[index]
    count = int(amount_list[index])
    rng = ws[f"A3:P{count + 2}"]
    for i, row_data in enumerate(rng):
        for index_cell, cell in enumerate(row_data):
            check_cell_value(index_cell, cell, temp)
        temp_list.append(
            [
                row_data[0].value,
                temp[0],
                row_data[3].value,
                row_data[4].value,
                row_data[7].value,
                row_data[8].value,
                temp[1],
                temp[2],
                row_data[13].value,
                temp[3],
                temp[4],
            ]
        )
        temp = []
    
    return temp_list

def write_to_csv(filename, fork_list):
    with open(filename, "w", newline="") as file:
        writer = csv.writer(file)
        for i, row in enumerate((fork_list)):
            if i == 0:
                continue
            writer.writerow(row)

def write_to_db(c, table_name, tuple_list):
    c.execute(
        f"CREATE TABLE IF NOT EXISTS {table_name}(id int, maker text, model text, serial_number text, height text, c_t text, attachment text, year text, hour_meter text, applicable text, amount text)"
    )
    for row in tuple_list:
        c.execute(
            f"insert into {table_name}(id, maker, model, serial_number, height, c_t, attachment, year, hour_meter, applicable, amount)"
            f"values {row}"
        )
    conn.commit()


def update_sheet(ws, rng, index, count, fork_list):
    for i, row in enumerate(fork_list):
        if i == 0:
            ws[rng[i][0].coordinate] = row[i]
            index += 1
            continue
        for i2, cell in enumerate(row):
            ws[rng[i][i2].coordinate] = cell
        index += 1
    return index

# initialization
translator = Translator()
hizuke = "20231117"
wbname = hizuke
sheetname = "total"
PRICE_DIR = "fobprice"
fob_file_name = f"{hizuke[0:4]}_{hizuke[4:6]}_{hizuke[6:]}.txt"
wb = openpyxl.load_workbook(f"stocklist/st_{wbname}.xlsx", data_only=True)
ws4 = wb.create_sheet(title=sheetname) if sheetname not in wb.sheetnames else print("シートは存在します")
ws1 = wb[sheetname]
sanichi = settings.TEST

# loading data from sanichi
for i, line in enumerate(sanichi):
    ws1[f"A{i + 1}"] = sanichi[i]

dt_now = datetime.datetime.now()
ws1["J3"] = dt_now.strftime("%Y/%m/%d")

rng1 = ws1[f"A{len(sanichi) + 1}:K{len(sanichi) + 1}"]
for i, head in enumerate(["id", "maker", "model", "serial_number", "height", "c_t", "attachment", "year", "hour_meter", "applicable", "fob_price"]):
    ws1[rng1[0][i].coordinate] = head

with open(path.join(PRICE_DIR, fob_file_name)) as f:
    amount_list = [x.rstrip() for x in f.readlines()]

fork_lists = [process_worksheet(wb, i, [], amount_list) for i in range(4)]
tuple_lists = [[tuple(["None" if v is None else v for v in row]) for row in fork_list] for fork_list in fork_lists]

# Save to csv
CSV_DIR = path.join("csv", f"{wbname}_c_t")
makedirs(CSV_DIR, exist_ok=True)

file_names = ['diesel_forklift.csv', 'gasoline_forklift.csv', 'battery_forklift.csv', 'shovelloader_forklift.csv']
for i, file_name in enumerate(file_names):
    write_to_csv(path.join(CSV_DIR, file_name), fork_lists[i])

# Save to database
SQLITE_DIR = "sqlite_c_t"
makedirs(SQLITE_DIR, exist_ok=True)

conn = sqlite3.connect(path.join(SQLITE_DIR, f"{wbname}.db"))
c = conn.cursor()
table_names = ["diesel_stocklist", "gasoline_stocklist", "battery_stocklist", "shovelloader_stocklist"]
for i, table_name in enumerate(table_names):
    write_to_db(c, table_name, tuple_lists[i])
conn.close()

# update workbook
names = ['diesel', 'gasoline', 'battery', 'shovelloader']
lastindex = len(sanichi) + 1
for i in range(4):
    rng = ws1[f"A{lastindex+1}:K{lastindex+1+int(amount_list[i])}"]
    lastindex = update_sheet(ws1, rng, lastindex, int(amount_list[i]), fork_lists[i])

wb.save(f"stocklist/st_{wbname}.xlsx")
