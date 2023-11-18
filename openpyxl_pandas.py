import pandas as pd
import openpyxl
from openpyxl.styles import Font
import numpy
import datetime
import re

st_name = "st_20220916"
forklift_csv_list = ["battery_forklift","diesel_forklift", "gasoline_forklift", "shovelloader_forklift"]
chassis_list = ["64359","F14F-11468","23199", "26038"]
data_list = []

for i, forklift in enumerate(forklift_csv_list):
    tmp = pd.read_csv(f"csv/{st_name}/{forklift_csv_list[i]}.csv", header=None)
    data_list.append(tmp)

df = pd.concat(data_list,ignore_index = True)

#ここのdf作成
diesel_df =  pd.read_csv(f"csv/{st_name}/{forklift_csv_list[1]}.csv", header=None)
gasoline_df =  pd.read_csv(f"csv/{st_name}/{forklift_csv_list[2]}.csv", header=None)
battery_df =  pd.read_csv(f"csv/{st_name}/{forklift_csv_list[0]}.csv", header=None)
shovelloader_df =  pd.read_csv(f"csv/{st_name}/{forklift_csv_list[3]}.csv", header=None)

diesel_selected_df = diesel_df[(diesel_df[3] == f"{chassis_list[0]}") | (diesel_df[3] == f"{chassis_list[1]}") | (diesel_df[3] == f"{chassis_list[2]}")]
""" | (diesel_df[3] == f"{chassis_list[3]}")"""

gasoline_selected_df = gasoline_df[(gasoline_df[3] == f"{chassis_list[0]}") | (gasoline_df[3] == f"{chassis_list[1]}") | (gasoline_df[3] == f"{chassis_list[2]}") | (gasoline_df[3] == f"{chassis_list[3]}")]#print(df.columns)

battery_selected_df = battery_df[(battery_df[3] == f"{chassis_list[0]}") | (battery_df[3] == f"{chassis_list[1]}") | (battery_df[3] == f"{chassis_list[2]}") | (battery_df[3] == f"{chassis_list[3]}")]

shovelloader_selected_df = shovelloader_df[(shovelloader_df[3] == f"{chassis_list[0]}") | (shovelloader_df[3] == f"{chassis_list[1]}") | (shovelloader_df[3] == f"{chassis_list[2]}") | (shovelloader_df[3] == f"{chassis_list[3]}")]

#print(gasoline_selected_df)
#print(diesel_selected_df)
#print(battery_selected_df)
#print(shovelloader_selected_df)

df_concat = pd.concat([diesel_selected_df,gasoline_selected_df,battery_selected_df])
selected_list = df_concat.values.tolist()
print(selected_list)

#diesel_list = diesel_selected_df.values.tolist()
#gasoline_list = gasoline_selected_df.values.tolist()
#battery_list = battery_selected_df.values.tolist()
#print(diesel_list)
#print(gasoline_list)
#print(battery_list)

#maker = df.loc[:,1]
#list_maker = maker.values.tolist()
#list_model = df.loc[:,2].values.tolist()
#list_chassis = df.loc[:,3].values.tolist()
#list_year = df.loc[:,6].values.tolist()
#list_hour = df.loc[:,7].values.tolist()
#list_height = df.loc[:,4].values.tolist()
#list_applicable = df.loc[:,8]
#list_amount = df.loc[:,9].values.tolist()

#print(list_maker)
#print(list_hour)
#print(list_chassis)
#print(list_height)
#print(list_model)
#print(list_year)
#print(type(maker))
pi_file = "3604C.xlsx"
wb_pi = openpyxl.load_workbook(f"PI/{pi_file}", data_only=True)
#ws2 = wb2.worksheets[0]
ws_pi = wb_pi["Sheet1"]

inv_file = "3604B.xlsx"
wb_inv = openpyxl.load_workbook(f"INV/{inv_file}", data_only=True)
ws_inv = wb_inv['Sheet1']

pl_file = "3604B.xlsx"
wb_pl = openpyxl.load_workbook(f"PL/{pl_file}", data_only=True)
ws_pl = wb_pl['Sheet1']

si_file = "3604(C).xlsx"
wb_si = openpyxl.load_workbook(f"SHIPPING_INSTRUCTION/{si_file}", data_only=True)
ws_si = wb_si['Sheet1']#ws = sb.worksheets[0]
#インデックス番号の確認
#wb.index(ws)

## シート名のリスト
#>>> wb.sheetnames
#['Sheet1', 'Sheet2', 'Sheet3']
#
## シート名の確認
#>>> ws.title
#'Sheet1'
#
## シート名の変更
#>>> ws.title = "SheetOne"
#>>> wb.sheetnames
#['SheetOne', 'Sheet2', 'Sheet3']
#
#>>> ws.title ="Sheet1"
#>>> wb.sheetnames
#['Sheet1', 'Sheet2', 'Sheet3']
#values = []
#addrs = []
#rng1 = ws_pi["A1:J17"]
#for index_ad,row in enumerate(rng1):
#    for index_value,cell in enumerate(row):
#        addrs.append(cell.coordinate)
#        values.append(cell.value)

#print(values)
#print(addrs)


sis_number = "NO SIS3604/22(D)"
bookingno = "11111"
ship = "wanhai"
voy = "S202"
ship_voy = f"Shipped per {ship} Voy.{voy}"
dt_now = datetime.datetime.now()
dt_year = str(dt_now.year)
dt_month = "0" + str(dt_now.month)
dt_hizuke = str(dt_now.strftime("Date: Osaka, %B %m/%Y"))
stock_count = len(selected_list)
stock_one = 5
unit = "1 unit"
kgs = "kgs"
m3 = "M3"
lc_number = "LCX1111"
sis_cell = "H5"
ship_cell = "C10"
hizuke_cell = "H7"
inv_unit_count_cell = "C9"


ws_si["A1"] =f"INVOCE {sis_number}"
ws_si["F2"] = f"Booking No. {bookingno}"
ws_si["A21"] = f"   {ship}                         {voy}"
ws_si["A35"] = str(stock_count) + " units"
ws_si["C40"] = f"     {lc_number}"
ws_pi[sis_cell] = sis_number
ws_inv[sis_cell] = sis_number
ws_pl[sis_cell] = sis_number
ws_pi[hizuke_cell] = dt_hizuke                
ws_inv[hizuke_cell] = dt_hizuke
ws_pl[hizuke_cell]  = dt_hizuke
ws_inv[inv_unit_count_cell] = re.sub(r'[0-9]',str(stock_count),ws_inv[inv_unit_count_cell].value)
ws_inv[ship_cell] = ship_voy

#forで回さなくても特定のセルしか変更しないのでやめる
#rng1 = ws_pi["AA:J17"]
#for index_ad,row in enumerate(rng1):
#    for index_value,cell in enumerate(row):
#        if index_ad == 4:
#            if index_value == 3:
#                ws_pi[f"{cell.coordinate}"] = "PROFOLMA INVOICE"
#                ws_inv[f"{cell.coordinate}"] = "INVOICE"
#                ws_pl[f"{cell.coordinate}"] = "PACKING LIST"
#            if index_value == 7:
#                ws_pi[f"{cell.coordinate}"] = sis_number
#                ws_inv[f"{cell.coordinate}"] = sis_number
#                ws_pl[f"{cell.coordinate}"] = sis_number
#        if index_ad == 6:
#            if index_value == 7:
#                ws_pi[f"{cell.coordinate}"] = dt_hizuke                
#                ws_inv[f"{cell.coordinate}"] = dt_hizuke
#                ws_pl[f"{cell.coordinate}"]  = dt_hizuke
#        if index_ad == 8:
#            if index_value == 2:
##                ws_inv[f"{cell.coordinate}"] = cell.value
# 
#for i in range(len(values)):
#    if i == 47:
#        ws_pi[f"{addrs[i]}"] = sis_number
#        ws_inv[f"{addrs[i]}"] = sis_number
#        ws_pl[f"{addrs[i]}"] = sis_number
# 
#        
#        continue
#    if i == 67:
#        ws_pi[f"{addrs[i]}"] = sis_number
#        ws_inv[f"{addrs[i]}"] = sis_number
#        ws_pl[f"{addrs[i]}"] = sis_number
#        continue
#    ws_pi[f"{addrs[i]}"] = values[i]


A_list = []
F_list = []
H_list = []
G_list = []
E_list = []
I_list = []
J_list = []
for i in range(stock_count*stock_one):
    A_list.append(str(f"A{i+18}"))

for i in range(3,100,5):
    F_list.append(str(f"F{i+18}"))
    H_list.append(str(f"H{i+18}"))
    E_list.append(str(f"E{i+18}"))
    G_list.append(str(f"G{i+18}"))
    I_list.append(str(f"I{i+18}"))
    J_list.append(str(f"J{i+18}"))

stockone_list = []
for i,st in enumerate(selected_list):
    stocklist_one =[f"{st[1]}",f"{i+1}) {st[4]}-Meter {st[5]}",f"Year {st[6]}/hour {st[7]}",f"Model {st[2]}     (S/No.{st[3]})",""]
    stockone_list.extend(stocklist_one)

for i in range(0,len(A_list),5):
        ws_pi[A_list[i]] = stockone_list[i]
        ws_pi[A_list[i+1]] = stockone_list[i+1]
        ws_pi[A_list[i+2]] = stockone_list[i+2]
        ws_pi[A_list[i+3]] = stockone_list[i+3]
        ws_pi[A_list[i+4]] = stockone_list[i+4]
        ws_inv[A_list[i]] = stockone_list[i]
        ws_inv[A_list[i+1]] = stockone_list[i+1]
        ws_inv[A_list[i+2]] = stockone_list[i+2]
        ws_inv[A_list[i+3]] = stockone_list[i+3]
        ws_inv[A_list[i+4]] = stockone_list[i+4]
        ws_pl[A_list[i]] = stockone_list[i]
        ws_pl[A_list[i+1]] = stockone_list[i+1]
        ws_pl[A_list[i+2]] = stockone_list[i+2]
        ws_pl[A_list[i+3]] = stockone_list[i+3]
        ws_pl[A_list[i+4]] = stockone_list[i+4]

pl_weight_m3_list = []

with open(f"pl_weight_m3/3604C.txt") as f:
    for line in f:
        pl_weight_m3_list.append([x.strip("\n") for x in line.split()])

for i,st in enumerate(selected_list):
    ws_pi[F_list[i]] = unit
    ws_pi[H_list[i]] = int(st[9].replace(",",""))
    ws_inv[F_list[i]] = unit
    ws_inv[H_list[i]] = int(st[9].replace(",",""))
    ws_pl[E_list[i]] = unit
    ws_pl[H_list[i]] = kgs
    ws_pl[J_list[i]] = m3

pl_total_weight = 0
pl_total_m3 = 0.0

for i,w_m3_list in enumerate(pl_weight_m3_list):
    tmp = int(w_m3_list[0].replace(",",""))
    tmp2 = float(w_m3_list[1])
    pl_total_weight+=tmp
    pl_total_m3+=tmp2
    ws_pl[G_list[i]] = tmp
    ws_pl[I_list[i]] = tmp2

ws_si["G39"] = pl_total_weight
ws_si["G40"] = pl_total_m3








total_unit_inv_pi_cell = "F71"
total_unit_pl_cell = "E71"
total_amount_cell = "H71"
kgs_total_cell = "G71"
m3_total_cell = "I71"


ws_pi[total_unit_inv_pi_cell] = str(stock_count) + " units"
ws_inv[total_unit_inv_pi_cell] = str(stock_count) + " units"
ws_pl[total_unit_pl_cell] = str(stock_count) + " units"
ws_pi[total_amount_cell] = "=SUM(H20:H70)"
ws_inv[total_amount_cell] = "=SUM(H20:H70)"
ws_pl[kgs_total_cell] = "=SUM(G20:G70)"
ws_pl[m3_total_cell]  = "=SUM(I20:I70)"


#上記同様直接載せるに対しての方が良さそうなので却下
#rng1 = ws_pi["A71:J79"]
#for index_ad,row in enumerate(rng1):
#    for index_value,cell in enumerate(row):
#        if index_ad == 0:
#            if index_value == 5:
#                ws_pi[f"{cell.coordinate}"] = str(stock_count) + " units"
#            if index_value == 7:
#                ws_pi[f"{cell.coordinate}"] = "=SUM(H20:H70)"
#        ws_pi[f"{cell.coordinate}"] = cell.value


wb_pi.save("3604pi.xlsx")
wb_inv.save("3604inv.xlsx")
wb_pl.save("3604pl.xlsx")
wb_si.save("3604si.xlsx")





